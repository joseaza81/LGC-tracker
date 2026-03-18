"""
LGC Transfer Tracker
====================
Fetches daily transfer data from the CER REC Registry public API and builds
a rolling 1-year summary of LGC transfers by holdings account, with inferred
parent entity affiliations.

API endpoint (no auth required):
  https://rec-registry.gov.au/rec-registry/app/api/public-register/certificate-actions?date=YYYY-MM-DD

Usage:
  # Fetch today's data and update the tracker
  python lgc_tracker.py

  # Fetch a specific date
  python lgc_tracker.py --date 2025-03-17

  # Backfill a date range (e.g. past year)
  python lgc_tracker.py --backfill --start 2025-03-18 --end 2026-03-18

  # Show summary table in terminal
  python lgc_tracker.py --summary

Requirements:
  pip install requests openpyxl pandas
"""

import argparse
import json
import os
import time
from datetime import date, datetime, timedelta
from pathlib import Path

import pandas as pd
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side)
from openpyxl.utils import get_column_letter

# ── Config ──────────────────────────────────────────────────────────────────
API_BASE   = "https://rec-registry.gov.au/rec-registry/app/api/public-register/certificate-actions"
DATA_FILE  = "lgc_transfer_data.json"     # raw daily cache
EXCEL_OUT  = "LGC_Account_Flow_Tracker.xlsx"
LOOKBACK_DAYS = 365

# ── Parent entity inference ──────────────────────────────────────────────────
# Maps substrings found in account names → inferred parent entity.
# Extend this list as you discover new accounts in the data.
PARENT_ENTITY_MAP = {
    # Retailers / liable entities
    "origin":           "Origin Energy",
    "agl":              "AGL Energy",
    "energyaustralia":  "EnergyAustralia (TRUenergy)",
    "energy australia": "EnergyAustralia (TRUenergy)",
    "alinta":           "Alinta Energy",
    "simply energy":    "Simply Energy (ENGIE)",
    "engie":            "ENGIE",
    "powershop":        "Powershop (Meridian Energy)",
    "meridian":         "Meridian Energy",
    "erm":              "ERM Power (Shell Energy)",
    "shell energy":     "Shell Energy Australia",
    "flow power":       "Flow Power",
    "momentum":         "Momentum Energy (Hydro Tasmania)",
    "hydro tasmania":   "Hydro Tasmania",
    "tas networks":     "TasNetworks",
    "ausgrid":          "Ausgrid",
    "endeavour":        "Endeavour Energy",
    "essential":        "Essential Energy",
    "ergon":            "Ergon Energy (Energy Queensland)",
    "energex":          "Energy Queensland",
    "powerlink":        "Powerlink Queensland",
    "transgrid":        "TransGrid",
    "aemo":             "AEMO",
    "synergy":          "Synergy (WA)",
    "horizon":          "Horizon Power (WA)",
    "aurora":           "Aurora Energy (TAS)",
    "actew":            "ActewAGL",
    "jemena":           "Jemena",
    "citipower":        "CitiPower (CKI Group)",
    "powercor":         "Powercor (CKI Group)",
    "united energy":    "United Energy (CKI Group)",
    "sp ausnet":        "AusNet Services",
    "ausnet":           "AusNet Services",
    # Generators / project developers
    "macquarie":        "Macquarie Group",
    "infigen":          "Infigen Energy (Copenhagen Infrastructure)",
    "neoen":            "Neoen",
    "nexgen":           "NexGen Energy",
    "pacific hydro":    "Pacific Hydro (CPPIB)",
    "tilt":             "Tilt Renewables",
    "goldwind":         "Goldwind Australia",
    "epuron":           "Epuron",
    "amp capital":      "AMP Capital",
    "amp energy":       "AMP Energy",
    "genex":            "Genex Power",
    "acciona":          "Acciona",
    "iberdrola":        "Iberdrola",
    "enel":             "Enel Green Power",
    "fru":              "First Solar / FRU",
    "boral":            "Boral",
    "rio tinto":        "Rio Tinto",
    "bhp":              "BHP",
    "woodside":         "Woodside",
    "santos":           "Santos",
    # Traders / intermediaries
    "cge":              "Clean Generation Energy",
    "renewable energy hub": "Renewable Energy Hub",
    "eco2":             "Eco2",
    "greenpower":       "GreenPower",
    "solstice":         "Solstice Energy",
    "greenedge":        "GreenEdge",
}

def infer_parent(account_name: str) -> str:
    """Attempt to infer parent entity from account name substrings."""
    lower = account_name.lower()
    for keyword, parent in PARENT_ENTITY_MAP.items():
        if keyword in lower:
            return parent
    return "Unknown / Independent"


# ── API fetch ────────────────────────────────────────────────────────────────
def fetch_day(query_date: date, retries: int = 3) -> list[dict]:
    date_str = query_date.strftime("%Y-%m-%d")
    url = f"{API_BASE}?date={date_str}"
    
    for attempt in range(1, retries + 1):
        try:
            resp = requests.get(url, timeout=30)
            resp.raise_for_status()
            data = resp.json()

            # API may return a dict with a list inside, or a bare list
            if isinstance(data, dict):
                records = data.get("data", data.get("results", data.get("certificateActions", [])))
            elif isinstance(data, list):
                records = data
            else:
                return []

            # Filter: only process dict items, LGC transfers only
          print(f"  Raw response sample: {str(records[:2])}")  
          transfers = [
                r for r in records
                if isinstance(r, dict)
                and r.get("certificateType") == "LGC"
                and r.get("actionType") in ("Transfer offer", "Transfer accept")
            ]
            return transfers

        except requests.RequestException as e:
            print(f"  [attempt {attempt}/{retries}] Error fetching {date_str}: {e}")
            if attempt < retries:
                time.sleep(2 ** attempt)
    return []


# ── Data cache ───────────────────────────────────────────────────────────────
def load_cache(path: str) -> dict:
    if Path(path).exists():
        with open(path, "r") as f:
            return json.load(f)
    return {}

def save_cache(cache: dict, path: str):
    with open(path, "w") as f:
        json.dump(cache, f, indent=2)


# ── Build transfer records ────────────────────────────────────────────────────
def parse_actions(actions: list[dict], action_date: date) -> list[dict]:
    """
    Convert raw API actions into flat transfer records.
    Each certificateRange entry within a Transfer accept = one record.
    """
    records = []
    for action in actions:
        if action.get("actionType") != "Transfer accept":
            continue
        completed = action.get("completedTime", "")
        for cr in action.get("certificateRanges", []):
            if cr.get("certificateType") != "LGC":
                continue
            start = int(cr.get("startSerialNumber", 0))
            end   = int(cr.get("endSerialNumber", 0))
            volume = max(0, end - start + 1)
            owner  = cr.get("ownerAccount", "Unknown")
            records.append({
                "date":              action_date.isoformat(),
                "completed_time":    completed,
                "owner_account":     owner,
                "owner_account_id":  cr.get("ownerAccountID", ""),
                "parent_entity":     infer_parent(owner),
                "accreditation":     cr.get("accreditationCode", ""),
                "generation_year":   cr.get("generationYear", ""),
                "generation_state":  cr.get("generationState", ""),
                "fuel_source":       cr.get("fuelSource", ""),
                "start_serial":      start,
                "end_serial":        end,
                "volume_lgc":        volume,
                "status":            cr.get("status", ""),
            })
    return records


# ── Summary builder ───────────────────────────────────────────────────────────
def build_summary(all_records: list[dict]) -> pd.DataFrame:
    """
    Summarise total LGCs received per (owner_account, parent_entity)
    with monthly breakdown.
    """
    if not all_records:
        return pd.DataFrame()
    
    df = pd.DataFrame(all_records)
    df["date"] = pd.to_datetime(df["date"])
    df["month"] = df["date"].dt.to_period("M").astype(str)

    summary = (
        df.groupby(["owner_account", "parent_entity"])["volume_lgc"]
        .sum()
        .reset_index()
        .sort_values("volume_lgc", ascending=False)
        .rename(columns={"volume_lgc": "total_lgcs_received"})
    )

    # Monthly pivot
    monthly = (
        df.groupby(["owner_account", "month"])["volume_lgc"]
        .sum()
        .unstack(fill_value=0)
    )
    summary = summary.merge(monthly, on="owner_account", how="left").fillna(0)

    return summary


# ── Excel writer ──────────────────────────────────────────────────────────────
GREEN_DARK  = "1A5276"
GREEN_MID   = "1E8449"
GREEN_LIGHT = "D5F5E3"
WHITE       = "FFFFFF"
TEXT_DARK   = "1C2833"

def _tb():
    s = Side(style="thin", color="BDC3C7")
    return Border(left=s, right=s, top=s, bottom=s)

def _hfill(col): return PatternFill("solid", fgColor=col)
def _hfont(col=WHITE, sz=11, bold=True): return Font(name="Arial", size=sz, bold=bold, color=col)
def _bfont(col=TEXT_DARK, sz=10): return Font(name="Arial", size=sz, color=col)
def _ctr(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _lft(): return Alignment(horizontal="left", vertical="center")

def _style_hdr(ws, row, ncols, bg=GREEN_DARK):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _hfill(bg); cell.font = _hfont(); cell.alignment = _ctr(); cell.border = _tb()

def _style_row(ws, row, ncols, alt=False):
    bg = GREEN_LIGHT if alt else WHITE
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _hfill(bg); cell.font = _bfont(); cell.alignment = _lft(); cell.border = _tb()


def write_excel(all_records: list[dict], summary_df: pd.DataFrame, output_path: str):
    wb = Workbook()

    # ── Sheet 1: Summary by Account ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Account Flow Summary"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    ws.merge_cells("A1:Z1")
    c = ws["A1"]
    c.value = "LGC MARKET — HOLDINGS ACCOUNT TRANSFER FLOW TRACKER  |  Rolling 12 Months"
    c.font  = _hfont(sz=14); c.fill = _hfill(GREEN_DARK); c.alignment = _ctr()
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:Z2")
    c = ws["A2"]
    c.value = f"Source: CER REC Registry Public API  |  Last updated: {date.today().isoformat()}  |  Action type: Transfer accept (LGC only)"
    c.font  = Font(name="Arial", size=9, italic=True, color=WHITE)
    c.fill  = _hfill(GREEN_MID); c.alignment = _ctr()
    ws.row_dimensions[2].height = 18

    if summary_df.empty:
        ws["A3"].value = "No data loaded yet. Run: python lgc_tracker.py --backfill"
        wb.save(output_path)
        return

    # Determine month columns
    fixed_cols = ["owner_account", "parent_entity", "total_lgcs_received"]
    month_cols = [c for c in summary_df.columns if c not in fixed_cols]

    headers = ["Registry Account Name", "Parent / Affiliated Entity", "Total LGCs Received"] + month_cols
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i).value = h
    _style_hdr(ws, 3, len(headers), bg=GREEN_MID)
    ws.row_dimensions[3].height = 28

    for r_idx, (_, row) in enumerate(summary_df.iterrows(), 4):
        alt = r_idx % 2 == 0
        vals = [row["owner_account"], row["parent_entity"], int(row["total_lgcs_received"])] + \
               [int(row[m]) if m in row and row[m] > 0 else "-" for m in month_cols]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = val
            cell.fill  = _hfill(GREEN_LIGHT if alt else WHITE)
            cell.font  = _bfont()
            cell.alignment = _ctr() if c_idx > 2 else _lft()
            cell.border = _tb()
            if isinstance(val, int) and val > 0:
                cell.number_format = "#,##0"

    # Column widths
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 20
    for i in range(4, len(headers)+1):
        ws.column_dimensions[get_column_letter(i)].width = 12

    # ── Sheet 2: Raw Transfer Log ────────────────────────────────────────────
    ws2 = wb.create_sheet("Raw Transfer Log")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A3"

    ws2.merge_cells("A1:L1")
    c = ws2["A1"]
    c.value = "RAW LGC TRANSFER LOG — All accepted transfers fetched from CER API"
    c.font  = _hfont(sz=12); c.fill = _hfill(GREEN_DARK); c.alignment = _ctr()
    ws2.row_dimensions[1].height = 28

    raw_headers = [
        "Date", "Registry Account Name", "Account ID", "Parent Entity",
        "Accreditation Code", "Gen Year", "State", "Fuel Source",
        "Start Serial", "End Serial", "Volume (LGCs)", "Status"
    ]
    for i, h in enumerate(raw_headers, 1):
        ws2.cell(row=2, column=i).value = h
    _style_hdr(ws2, 2, len(raw_headers), bg=GREEN_MID)

    for r_idx, rec in enumerate(sorted(all_records, key=lambda x: x["date"], reverse=True), 3):
        alt = r_idx % 2 == 0
        vals = [
            rec["date"], rec["owner_account"], rec["owner_account_id"],
            rec["parent_entity"], rec["accreditation"], rec["generation_year"],
            rec["generation_state"], rec["fuel_source"],
            rec["start_serial"], rec["end_serial"], rec["volume_lgc"], rec["status"]
        ]
        for c_idx, val in enumerate(vals, 1):
            cell = ws2.cell(row=r_idx, column=c_idx)
            cell.value = val
            cell.fill  = _hfill(GREEN_LIGHT if alt else WHITE)
            cell.font  = _bfont()
            cell.alignment = _ctr() if c_idx not in [2,4,5,8,12] else _lft()
            cell.border = _tb()
        ws2.cell(row=r_idx, column=11).number_format = "#,##0"

    raw_widths = [12, 38, 14, 30, 20, 10, 8, 20, 16, 16, 16, 22]
    for i, w in enumerate(raw_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: Entity Affiliation Map ─────────────────────────────────────
    ws3 = wb.create_sheet("Entity Affiliation Map")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:D1")
    c = ws3["A1"]
    c.value = "PARENT ENTITY AFFILIATION MAP — Edit to refine inferences"
    c.font  = _hfont(sz=12); c.fill = _hfill(GREEN_DARK); c.alignment = _ctr()
    ws3.row_dimensions[1].height = 28

    for i, h in enumerate(["Keyword (in account name)", "Mapped Parent Entity", "Notes"], 1):
        ws3.cell(row=2, column=i).value = h
    _style_hdr(ws3, 2, 3, bg=GREEN_MID)

    for r_idx, (kw, parent) in enumerate(PARENT_ENTITY_MAP.items(), 3):
        ws3.cell(row=r_idx, column=1).value = kw
        ws3.cell(row=r_idx, column=2).value = parent
        ws3.cell(row=r_idx, column=3).value = ""
        _style_row(ws3, r_idx, 3, alt=r_idx%2==0)

    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 38
    ws3.column_dimensions["C"].width = 30

    wb.save(output_path)
    print(f"  ✔  Saved: {output_path}")


# ── CLI ───────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="LGC Transfer Flow Tracker")
    parser.add_argument("--date",     help="Fetch a single date YYYY-MM-DD")
    parser.add_argument("--backfill", action="store_true",
                        help="Fetch all missing dates in the rolling 1-year window")
    parser.add_argument("--start",    help="Backfill start date YYYY-MM-DD (default: 1 year ago)")
    parser.add_argument("--end",      help="Backfill end date YYYY-MM-DD (default: yesterday)")
    parser.add_argument("--summary",  action="store_true", help="Print summary table to terminal")
    parser.add_argument("--data",     default=DATA_FILE,  help="Path to JSON cache file")
    parser.add_argument("--output",   default=EXCEL_OUT,  help="Path to output Excel file")
    args = parser.parse_args()

    cache = load_cache(args.data)
    yesterday = date.today() - timedelta(days=1)  # API has 1-day latency

    # ── Determine which dates to fetch ──────────────────────────────────────
    dates_to_fetch: list[date] = []

    if args.date:
        dates_to_fetch = [datetime.strptime(args.date, "%Y-%m-%d").date()]

    elif args.backfill:
        start = datetime.strptime(args.start, "%Y-%m-%d").date() if args.start \
                else yesterday - timedelta(days=LOOKBACK_DAYS)
        end   = datetime.strptime(args.end,   "%Y-%m-%d").date() if args.end \
                else yesterday
        d = start
        while d <= end:
            if d.isoformat() not in cache:
                dates_to_fetch.append(d)
            d += timedelta(days=1)
        print(f"  Backfill: {len(dates_to_fetch)} missing dates "
              f"({start.isoformat()} → {end.isoformat()})")

    else:
        # Default: fetch yesterday if not already cached
        if yesterday.isoformat() not in cache:
            dates_to_fetch = [yesterday]
        else:
            print(f"  ✔  {yesterday.isoformat()} already in cache. "
                  "Use --backfill to refresh missing dates.")

    # ── Fetch & cache ─────────────────────────────────────────────────────────
    for d in dates_to_fetch:
        print(f"  Fetching {d.isoformat()}...", end=" ")
        actions  = fetch_day(d)
        records  = parse_actions(actions, d)
        cache[d.isoformat()] = records
        print(f"{len(records)} LGC transfers")
        time.sleep(0.5)  # be polite to the API

    save_cache(cache, args.data)

    # ── Prune cache to rolling 1-year window ─────────────────────────────────
    cutoff = (yesterday - timedelta(days=LOOKBACK_DAYS)).isoformat()
    cache  = {k: v for k, v in cache.items() if k >= cutoff}
    save_cache(cache, args.data)

    # ── Build summary & write Excel ───────────────────────────────────────────
    all_records = [rec for recs in cache.values() for rec in recs]
    print(f"\n  Total records in rolling window: {len(all_records)}")

    summary_df = build_summary(all_records)

    if args.summary and not summary_df.empty:
        print("\n─── Account Flow Summary (top 20) ───────────────────────────")
        print(summary_df[["owner_account", "parent_entity", "total_lgcs_received"]]
              .head(20).to_string(index=False))
        print()

    write_excel(all_records, summary_df, args.output)
    print(f"\n  Done. Open {args.output} to review.\n")


if __name__ == "__main__":
    main()
